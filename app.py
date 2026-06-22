import streamlit as st
import pandas as pd
import re
import io
import zipfile
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
st.set_page_config(page_title="PUMA Voucher SKU Tool", page_icon="🏷️", layout="wide")

# ─────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────

REGION_MARKETPLACES = {
    "PH": ["Lazada", "Shopee", "Zalora"],
    "MY": ["Lazada", "Shopee", "Zalora", "TikTok"],
    "SG": ["Lazada", "Shopee", "Zalora"],
}

REGION_CONFIG = {
    "PH": {
        "zecom_sheet": "PH", "zecom_read": "ph", "article_col": "PIM Article#",
        "threshold": 650, "currency": "PHP 650",
        "mp_flags": {"Lazada": "LAZADA", "Shopee": "SHOPEE", "Zalora": "ZALORA"},
        "default_excl": 71, "default_rrp": 32, "default_srp": 50,
    },
    "MY": {
        "zecom_sheet": "MY", "zecom_read": "header3", "article_col": "Style#",
        "threshold": 36, "currency": "RM 36",
        "mp_flags": {"Lazada": "Lazada", "Shopee": "Shopee",
                     "Zalora": "Zalora MP", "TikTok": "TIKTOK"},
        "default_excl": 51, "default_rrp": 27, "default_srp": 49,
    },
    "SG": {
        "zecom_sheet": "SG", "zecom_read": "header3", "article_col": "STYLE#",
        "threshold": 16, "currency": "SGD 16",
        "mp_flags": {"Lazada": "Lazada", "Shopee": "Shopee", "Zalora": "Zalora"},
        "default_excl": 52, "default_rrp": 26, "default_srp": 50,
    },
}

PID_MARKETPLACES = {"Shopee", "TikTok"}   # marketplaces that group variants under a Product ID

# ─────────────────────────────────────────────────────────────────
# ZECOM READING & VALIDATION
# ─────────────────────────────────────────────────────────────────

def validate_zecom_region(file_bytes: bytes, selected_region: str):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return False, f"Cannot read ZeCom file: {e}"
    if selected_region == "PH":
        if "PH" not in sheets:
            extra = " (looks like MY/SG tracker)" if ("MY" in sheets or "SG" in sheets) else ""
            return False, f"⚠️ Wrong file — selected **PH** but 'PH' sheet not found{extra}."
    else:
        if selected_region not in sheets:
            extra = " (looks like PH tracker)" if "PH" in sheets else ""
            return False, f"⚠️ Wrong file — selected **{selected_region}** but sheet not found{extra}."
    return True, "OK"


@st.cache_data(show_spinner=False)
def read_zecom(file_bytes: bytes, region: str) -> pd.DataFrame:
    cfg = REGION_CONFIG[region]
    if cfg["zecom_read"] == "ph":
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="PH", header=1)
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=cfg["zecom_sheet"], header=3)
    return df


# ─────────────────────────────────────────────────────────────────
# COLUMN DETECTION HELPERS
# ─────────────────────────────────────────────────────────────────

_EXCL_RE = re.compile(
    r'open for all|exclude|vc only|vc max|vc -|shopee exclusive|platform vc', re.I
)

def excel_col_letter(idx: int) -> str:
    """0-based column index -> Excel column letter (0->A, 1->B, ..., 25->Z, 26->AA, ...)."""
    idx += 1
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters

def col_options(df):
    """Labels like 'BT: Exclusion' — matches the column letters seen in the actual Excel file."""
    return [f"{excel_col_letter(i)}: {col}" for i, col in enumerate(df.columns)]

def _col_score_excl(series):
    vals = series.dropna().astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty: return 0
    return int(vals.apply(lambda v: bool(_EXCL_RE.search(v))).sum())

def _col_by_name(df, keywords):
    for kw in keywords:
        for i, col in enumerate(df.columns):
            if kw.lower() in str(col).lower():
                return i
    return None

def guess_excl_idx(df, fallback):
    best_col, best_score = fallback, 0
    for i in range(len(df.columns)):
        s = _col_score_excl(df.iloc[:, i])
        if s > best_score:
            best_score, best_col = s, i
    return best_col

def guess_rrp_idx(df, fallback):
    idx = _col_by_name(df, ["rrp"])
    return idx if idx is not None else fallback

def guess_srp_idx(df, fallback):
    for kw in ["srp ao", "ec srp", "srp"]:
        for i, col in enumerate(df.columns):
            if kw in str(col).lower() and "rrp" not in str(col).lower():
                return i
    return fallback

def sample_vals(df, col_idx, n=6):
    vals = df.iloc[:, col_idx].dropna().unique()[:n]
    return ", ".join(str(v) for v in vals) if len(vals) else "(no values)"

def get_unique_remarks(df, col_idx):
    vals = df.iloc[:, col_idx].dropna().astype(str).str.strip()
    vals = vals[(vals != "") & (vals.str.lower() != "nan")]
    if vals.empty: return []
    return vals.value_counts().index.tolist()


# ─────────────────────────────────────────────────────────────────
# INVENTORY READING
# ─────────────────────────────────────────────────────────────────

def _normalize_inv(df, ean_c, stock_c):
    ec = next((c for c in ean_c   if c in df.columns), None)
    sc = next((c for c in stock_c if c in df.columns), None)
    if ec is None or sc is None: return None
    out = df[[ec, sc]].copy()
    out.columns = ["EAN", "Stock"]
    out["EAN"]   = out["EAN"].astype(str).str.strip()
    out["Stock"] = pd.to_numeric(out["Stock"], errors="coerce").fillna(0)
    return out[out["EAN"].str.match(r"^\d{13}$")]

def read_inventory(file_bytes, filename):
    ec = ["EAN", "Sku", "PROD_CODE", "SellerSku"]
    sc = ["Avail_Qty", "QtyAvailable", "QTY", "Quantity"]
    if filename.lower().endswith(".csv"):
        return _normalize_inv(pd.read_csv(io.BytesIO(file_bytes)), ec, sc)
    r = _normalize_inv(pd.read_excel(io.BytesIO(file_bytes)), ec, sc)
    return r if r is not None else _normalize_inv(
        pd.read_excel(io.BytesIO(file_bytes), header=4), ec, sc)


# ─────────────────────────────────────────────────────────────────
# SPECIAL ARTICLE EXCLUSION PARSING
# ─────────────────────────────────────────────────────────────────

def parse_special_articles(text_input: str, file_bytes: bytes = None, filename: str = None) -> set:
    """Combine articles typed in the text box with articles from an uploaded file."""
    articles = set()

    if text_input and text_input.strip():
        for tok in re.split(r"[,\n\r\t]+", text_input.strip()):
            tok = tok.strip()
            if tok:
                articles.add(tok)

    if file_bytes and filename:
        try:
            if filename.lower().endswith(".csv"):
                fdf = pd.read_csv(io.BytesIO(file_bytes), header=None)
            else:
                fdf = pd.read_excel(io.BytesIO(file_bytes), header=None)
            # Try to detect a header row — if first cell of col0 looks like text, skip it
            first_col = fdf.iloc[:, 0].dropna().astype(str).str.strip()
            for v in first_col:
                if v and v.lower() not in ("article", "article no", "article number", "sku", "nan"):
                    articles.add(v)
        except Exception:
            pass

    return articles


# ─────────────────────────────────────────────────────────────────
# ZECOM ARTICLE-LEVEL PROCESSING  (status + price + remark + special)
# ─────────────────────────────────────────────────────────────────

def classify_row(article, mp_status, status_ok, price_ok, remark,
                 eligible_remarks, include_no_remark, special_articles):
    """
    Priority order: Special Article > MP Status > Price > Remark.
    Returns (status, reason) where status in {'eligible','ineligible','no_remark'}.
    """
    if article in special_articles:
        return "ineligible", "Special Article Exclusion"

    if not status_ok:
        disp = mp_status if mp_status not in ("", "NAN", None) else "BLANK"
        return "ineligible", f"MP Status = {disp}"

    if not price_ok:
        return "ineligible", "Price below threshold (RRP/SRP)"

    if pd.isna(remark) or str(remark).strip() == "" or str(remark).strip().lower() == "nan":
        if include_no_remark:
            return "eligible", ""
        return "no_remark", "No remark (not included)"

    r = str(remark).strip()
    if r in eligible_remarks:
        return "eligible", ""
    return "ineligible", f'Remark not selected ("{r}")'


def process_zecom(zecom_df, region, marketplace, excl_idx, rrp_idx, srp_idx,
                  eligible_remarks: set, include_no_remark: bool,
                  special_articles: set) -> pd.DataFrame:
    """
    Returns per-article DataFrame:
      [article, mp_status, rrp, srp, remark, status, reason]
    status: 'eligible' | 'ineligible' | 'no_remark'
    """
    cfg = REGION_CONFIG[region]
    df  = zecom_df.copy()

    # MP status — kept at row level (NOT dropped), so a "NO" article
    # can still trigger Product-ID-level exclusion downstream.
    mp_col = cfg["mp_flags"].get(marketplace)
    if mp_col and mp_col in df.columns:
        mp_status_disp = df[mp_col].fillna("").astype(str).str.strip().str.upper()
        status_ok = mp_status_disp == "YES"
    else:
        mp_status_disp = pd.Series(["N/A"] * len(df), index=df.index)
        status_ok = pd.Series([True] * len(df), index=df.index)

    threshold = cfg["threshold"]
    rrp = pd.to_numeric(df.iloc[:, rrp_idx], errors="coerce")
    srp = pd.to_numeric(df.iloc[:, srp_idx], errors="coerce")
    srp_ok   = (srp == 0) | (srp >= threshold)
    price_ok = (rrp > threshold) & srp_ok

    remark_vals  = df.iloc[:, excl_idx]
    article_vals = df[cfg["article_col"]].astype(str).str.strip()

    work = pd.DataFrame({
        "article":   article_vals.values,
        "mp_status": mp_status_disp.values,
        "status_ok": status_ok.values,
        "rrp":       rrp.values,
        "srp":       srp.values,
        "price_ok":  price_ok.values,
        "remark":    remark_vals.values,
    })

    work = work[work["article"].str.match(r"^[\w_\-]+$", na=False)]
    work = work[work["article"].str.lower() != "nan"]
    work = work.drop_duplicates(subset=["article"])

    statuses, reasons = [], []
    for row in work.itertuples(index=False):
        s, r = classify_row(row.article, row.mp_status, row.status_ok, row.price_ok,
                            row.remark, eligible_remarks, include_no_remark, special_articles)
        statuses.append(s); reasons.append(r)
    work["status"] = statuses
    work["reason"] = reasons

    return work[["article", "mp_status", "rrp", "srp", "remark", "status", "reason"]]


# ─────────────────────────────────────────────────────────────────
# EAN MAPPING & ELIGIBILITY SETS
# ─────────────────────────────────────────────────────────────────

def map_to_eans(article_df, content_df, inventory_df):
    merged = article_df.merge(
        content_df.rename(columns={"Color_No": "article"}), on="article", how="inner"
    )
    merged["EAN"] = merged["EAN"].astype(str).str.strip()
    merged = merged.merge(inventory_df.rename(columns={"Stock": "stock_qty"}),
                          on="EAN", how="left")
    merged["stock_qty"] = merged["stock_qty"].fillna(0)
    merged["has_stock"] = merged["stock_qty"] > 0
    return merged[["article", "EAN", "mp_status", "rrp", "srp", "remark",
                   "status", "reason", "stock_qty", "has_stock"]]

def eligible_ean_set(df): return set(df[(df["status"] == "eligible") & df["has_stock"]]["EAN"])
def excluded_ean_set(df): return set(df[df["status"] == "ineligible"]["EAN"])
def no_remark_ean_set(df): return set(df[df["status"] == "no_remark"]["EAN"])


# ─────────────────────────────────────────────────────────────────
# MARKETPLACE PROCESSORS
# ─────────────────────────────────────────────────────────────────

def process_lazada(ean_df, lazada_bytes):
    df   = pd.read_excel(io.BytesIO(lazada_bytes), sheet_name="template", header=0)
    data = df.iloc[3:].copy(); data.columns = df.columns
    active = data[data["status"].astype(str).str.lower() == "active"].copy()
    active["_ean"] = active["SellerSKU"].astype(str).str.strip()
    ok = eligible_ean_set(ean_df)
    skus = active[active["_ean"].isin(ok)]["Shop SKU"].dropna().apply(clean_id_str)
    return skus.dropna().unique().tolist()


def _read_shopee_zip(zip_bytes):
    dfs = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = sorted(n for n in zf.namelist() if n.endswith(".xlsx"))
        bar = st.progress(0, text="Reading Shopee export files…")
        for i, name in enumerate(names):
            with zf.open(name) as f:
                dfs.append(pd.read_excel(f, engine="calamine", header=2, skiprows=[3, 4]))
            bar.progress((i + 1) / len(names), text=f"Reading Shopee file {i+1}/{len(names)}…")
        bar.empty()
    return pd.concat(dfs, ignore_index=True)


def clean_id_str(val):
    """
    Clean an ID value (Product ID, Shop SKU, etc.) for exact output.
    Handles the common pandas pitfall where a numeric ID column containing
    even a single blank cell gets promoted to float64, so whole-number IDs
    come out as '18890032587.0' instead of '18890032587'.
    Already-clean strings pass through untouched.
    """
    if pd.isna(val):
        return None
    if isinstance(val, float):
        return str(int(val)) if val.is_integer() else str(val)
    s = str(val).strip()
    if re.match(r"^-?\d+\.0+$", s):       # defensive: catches '123.0' even if already stringified upstream
        s = s.split(".")[0]
    return s


def _extract_ean(sku_val, parent_val):
    for v in (sku_val, parent_val):
        if pd.notna(v):
            try:
                s = str(int(float(v)))
                if re.match(r"^\d{13}$", s): return s
            except (ValueError, TypeError):
                s = str(v).strip()
                if re.match(r"^\d{13}$", s): return s
    return None


def build_pid_decisions(combined_df, ean_df):
    """
    PID-level eligibility decision for Shopee/TikTok.
    combined_df must have '_ean' and '_pid' columns already populated.
    A PID is EXCLUDED if ANY variant EAN is in the excluded set
    (covers: ineligible remark, price fail, MP status = NO, special-article exclusion).
    A PID is INCLUDED only if it has zero excluded variants AND at least one eligible-in-stock variant.
    """
    ok_eans   = eligible_ean_set(ean_df)
    excl_eans = excluded_ean_set(ean_df)
    ean_info  = ean_df.drop_duplicates(subset=["EAN"]).set_index("EAN")[["article", "reason"]].to_dict("index")

    decisions = {}
    for pid, grp in combined_df.groupby("_pid"):
        eans = set(grp["_ean"].dropna())
        excl_hits = eans & excl_eans
        ok_hits   = eans & ok_eans

        if excl_hits:
            reasons = []
            for e in excl_hits:
                info = ean_info.get(e)
                if info:
                    reasons.append(f"{info['article']} ({e}): {info['reason']}")
                else:
                    reasons.append(f"EAN {e}: excluded")
            decisions[pid] = {
                "decision": "Excluded",
                "reason": "; ".join(reasons),
                "total_variants": len(eans), "eligible_variants": len(ok_hits),
                "excluded_variants": len(excl_hits),
            }
        elif ok_hits:
            decisions[pid] = {
                "decision": "Included",
                "reason": f"{len(ok_hits)} eligible variant(s) in stock",
                "total_variants": len(eans), "eligible_variants": len(ok_hits),
                "excluded_variants": 0,
            }
        else:
            decisions[pid] = {
                "decision": "Excluded",
                "reason": "No eligible-in-stock variant found",
                "total_variants": len(eans), "eligible_variants": 0,
                "excluded_variants": 0,
            }
    return decisions


def process_shopee(ean_df, zip_bytes):
    combined = _read_shopee_zip(zip_bytes)
    combined["_ean"] = combined.apply(lambda r: _extract_ean(r.get("SKU"), r.get("Parent SKU")), axis=1)
    combined["_pid"] = combined["Product ID"].apply(clean_id_str)
    decisions = build_pid_decisions(combined, ean_df)
    ids = [pid for pid, d in decisions.items() if d["decision"] == "Included"]
    return ids, decisions


def process_zalora(ean_df, eligible_bytes, content_df):
    df = pd.read_excel(io.BytesIO(eligible_bytes), sheet_name="Eligible Products")
    ok = eligible_ean_set(ean_df); nr = no_remark_ean_set(ean_df)
    ean2art = dict(zip(content_df["EAN"].astype(str).str.strip(),
                       content_df["Color_No"].astype(str).str.strip()))
    df["_ean"]       = df["Seller SKU"].astype(str).str.strip()
    df["Article No"] = df["_ean"].map(ean2art)
    df["Voucher Eligible"] = df["_ean"].apply(
        lambda e: "Yes" if e in ok else ("No Remark" if e in nr else "No"))
    df.drop(columns=["_ean"], inplace=True)
    return df


def process_tiktok(ean_df, tiktok_bytes):
    df = pd.read_excel(io.BytesIO(tiktok_bytes), sheet_name="Template", header=2, skiprows=[3, 4])
    df["_ean"] = df["Seller SKU"].apply(lambda v: _extract_ean(v, None))
    df["_pid"] = df["Product ID"].apply(clean_id_str)
    decisions = build_pid_decisions(df, ean_df)
    ids = [pid for pid, d in decisions.items() if d["decision"] == "Included"]
    return ids, decisions


# ─────────────────────────────────────────────────────────────────
# OUTPUT HELPERS
# ─────────────────────────────────────────────────────────────────

def _to_excel_multi(sheets: dict):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for name, d in sheets.items():
            d.to_excel(w, index=False, sheet_name=name[:31])
    return buf.getvalue()

def make_lazada_output(ids): return _to_excel_multi({"Sheet1": pd.DataFrame({"SHOP SKU": ids})})
def make_shopee_output(ids): return _to_excel_multi({"Sheet1": pd.DataFrame({"Product ID": ids})})
def make_zalora_output(ann): return _to_excel_multi({"Eligible Products": ann})


def make_summary_excel(ean_df, region, marketplace, pct, voucher_type, pid_decisions=None):
    detail = ean_df.copy()
    detail["status"] = detail["status"].map(
        {"eligible": "Eligible", "ineligible": "Ineligible", "no_remark": "No Remark"})
    detail = detail.rename(columns={
        "article": "Article", "mp_status": "MP Status", "rrp": "RRP", "srp": "SRP",
        "remark": "Remark", "status": "Status", "reason": "Exclusion Reason",
        "stock_qty": "Stock Qty", "has_stock": "In Stock",
    })
    detail.insert(0, "Region", region)
    detail.insert(1, "Marketplace", marketplace)
    detail.insert(2, "Voucher %", pct)
    detail.insert(3, "Voucher Type", voucher_type)
    cols = ["Region", "Marketplace", "Voucher %", "Voucher Type", "Article", "EAN",
            "MP Status", "RRP", "SRP", "Remark", "Status", "Exclusion Reason",
            "Stock Qty", "In Stock"]
    detail = detail[[c for c in cols if c in detail.columns]]

    sheets = {"Article_EAN_Detail": detail}

    total    = len(detail)
    n_elig   = (detail["Status"] == "Eligible").sum()
    n_inelig = (detail["Status"] == "Ineligible").sum()
    n_norem  = (detail["Status"] == "No Remark").sum()
    n_special = (detail["Exclusion Reason"] == "Special Article Exclusion").sum()
    n_price   = detail["Exclusion Reason"].astype(str).str.startswith("Price below threshold").sum()
    n_status  = detail["Exclusion Reason"].astype(str).str.startswith("MP Status").sum()
    n_remark  = detail["Exclusion Reason"].astype(str).str.startswith("Remark not selected").sum()

    stats_rows = [
        ("Region", region), ("Marketplace", marketplace),
        ("Voucher %", pct), ("Voucher Type", voucher_type),
        ("Total Article-EAN rows", total),
        ("Eligible", int(n_elig)), ("Ineligible", int(n_inelig)), ("No Remark", int(n_norem)),
        ("Excluded — Special Article", int(n_special)),
        ("Excluded — Price Below Threshold", int(n_price)),
        ("Excluded — MP Status Not YES", int(n_status)),
        ("Excluded — Remark Not Selected", int(n_remark)),
    ]

    if pid_decisions:
        pid_rows = [{
            "Product ID": pid, "Decision": d["decision"],
            "Total Variants": d["total_variants"],
            "Eligible Variants": d["eligible_variants"],
            "Excluded Variants": d["excluded_variants"],
            "Reason": d["reason"],
        } for pid, d in pid_decisions.items()]
        sheets["Product_ID_Summary"] = pd.DataFrame(pid_rows)

        n_pid_total = len(pid_decisions)
        n_pid_incl  = sum(1 for d in pid_decisions.values() if d["decision"] == "Included")
        stats_rows += [
            ("Total Product IDs", n_pid_total),
            ("Included Product IDs", n_pid_incl),
            ("Excluded Product IDs", n_pid_total - n_pid_incl),
        ]

    sheets["Summary_Stats"] = pd.DataFrame(stats_rows, columns=["Metric", "Value"])
    return _to_excel_multi(sheets)


# ─────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────

def main():
    st.title("🏷️ PUMA Voucher Eligible SKU Tool")
    st.caption("Generate marketplace-ready voucher eligible SKU lists from ZeCom tracker data.")

    # ── ① REGION & MARKETPLACE ───────────────────────────────────
    st.markdown("---")
    st.subheader("① Region & Marketplace")
    c1, c2 = st.columns(2)
    with c1: region      = st.selectbox("Region", ["PH", "MY", "SG"])
    with c2: marketplace = st.selectbox("Marketplace", REGION_MARKETPLACES[region])

    # ── ② UPLOAD FILES ────────────────────────────────────────────
    st.markdown("---")
    st.subheader("② Upload Files")
    cf1, cf2 = st.columns(2)
    with cf1:
        st.markdown("**Core files**")
        zecom_file   = st.file_uploader(
            f"ZeCom Tracker ({'PH file' if region == 'PH' else 'MY + SG file'})",
            type=["xlsx"], key="zecom")
        content_file = st.file_uploader("Content File", type=["xlsx"], key="content")
        inv_file     = st.file_uploader(f"Inventory File ({region})", type=["xlsx", "csv"], key="inv")
    with cf2:
        st.markdown(f"**{marketplace} export**")
        if   marketplace == "Lazada": mp_file = st.file_uploader("Lazada Product Export (.xlsx)",          type=["xlsx"], key="mp")
        elif marketplace == "Shopee": mp_file = st.file_uploader("Shopee Export ZIP (.zip — all batches)", type=["zip"],  key="mp")
        elif marketplace == "Zalora": mp_file = st.file_uploader("Zalora EligibleProducts File (.xlsx)",   type=["xlsx"], key="mp")
        elif marketplace == "TikTok": mp_file = st.file_uploader("TikTok Seller Center Export (.xlsx)",    type=["xlsx"], key="mp")
        else: mp_file = None

    # ── ②b SPECIAL ARTICLE EXCLUSION ─────────────────────────────
    st.markdown("---")
    st.subheader("②b  Special Article Exclusion  (optional)")
    st.caption(
        "Articles entered here are **always excluded**, on every channel, before any other "
        "eligibility check. For Shopee/TikTok, if a special article belongs to a Product ID, "
        "the **entire Product ID** is excluded."
    )
    sa1, sa2 = st.columns(2)
    with sa1:
        special_text = st.text_area(
            "Paste article numbers (comma or newline separated)",
            placeholder="521351_01\n521352_02, 521353_03",
            height=100,
        )
    with sa2:
        special_file = st.file_uploader(
            "...or upload a file (first column = article numbers)",
            type=["csv", "xlsx"], key="special_file",
        )

    special_articles = parse_special_articles(
        special_text,
        special_file.getvalue() if special_file else None,
        special_file.name if special_file else None,
    )
    if special_articles:
        st.info(f"🚫 **{len(special_articles)}** special article(s) will always be excluded.")

    # ── ③ ZECOM COLUMNS + REMARKS + VOUCHER ──────────────────────
    excl_idx = rrp_idx = srp_idx = zecom_df = None
    voucher_configs = []
    voucher_type = "Regular VC"

    if zecom_file:
        ok_region, err_msg = validate_zecom_region(zecom_file.getvalue(), region)
        if not ok_region:
            st.error(err_msg); st.stop()

        st.markdown("---")
        st.subheader("③ Select ZeCom Columns")
        st.caption("The app pre-selects the most likely columns. Check sample values and adjust if needed.")

        with st.spinner("Reading ZeCom file…"):
            zecom_df = read_zecom(zecom_file.getvalue(), region)

        cfg    = REGION_CONFIG[region]
        opts   = col_options(zecom_df)
        n_cols = len(zecom_df.columns)
        safe   = lambda i: min(i, n_cols - 1)

        d_excl = safe(guess_excl_idx(zecom_df, cfg["default_excl"]))
        d_rrp  = safe(guess_rrp_idx (zecom_df, cfg["default_rrp"]))
        d_srp  = safe(guess_srp_idx (zecom_df, cfg["default_srp"]))

        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            st.markdown("**📋 Exclusion / Campaign Column**")
            excl_sel = st.selectbox("excl", opts, index=d_excl, key="sel_excl", label_visibility="collapsed")
            excl_idx = opts.index(excl_sel)
            st.caption(f"Sample: `{sample_vals(zecom_df, excl_idx)}`")
        with sc2:
            st.markdown("**💰 RRP Column**")
            rrp_sel = st.selectbox("rrp", opts, index=d_rrp, key="sel_rrp", label_visibility="collapsed")
            rrp_idx = opts.index(rrp_sel)
            st.caption(f"Sample: `{sample_vals(zecom_df, rrp_idx)}`")
        with sc3:
            st.markdown("**🏷️ SRP Column**")
            srp_sel = st.selectbox("srp", opts, index=d_srp, key="sel_srp", label_visibility="collapsed")
            srp_idx = opts.index(srp_sel)
            st.caption(f"Sample: `{sample_vals(zecom_df, srp_idx)}`")

        cfg_cur = cfg["currency"]
        st.info(f"**Price filter:** RRP > {cfg_cur}  |  SRP = 0 (full price ✓) or SRP ≥ {cfg_cur}  |  "
                f"SRP > 0 but < {cfg_cur} → excluded")

        # ── ④ VOUCHER CONFIGURATION (multiple vouchers, each with its own remarks) ──
        st.markdown("---")
        st.subheader("④ Voucher Configuration")
        st.caption("Add one row per voucher % you want to generate this run — "
                  "each row has its own remarks selection.")

        unique_remarks = get_unique_remarks(zecom_df, excl_idx)

        voucher_type = st.radio("Voucher Type (applies to all vouchers below)",
                                ["Regular VC", "Bundle Discount"], horizontal=True)

        if "voucher_row_ids" not in st.session_state:
            st.session_state.voucher_row_ids = [0]
        if "voucher_row_counter" not in st.session_state:
            st.session_state.voucher_row_counter = 1

        mp_col_for_preview = cfg["mp_flags"].get(marketplace)

        def _render_voucher_row(rid, position):
            st.markdown(f"**Voucher {position}**")
            rcol1, rcol2, rcol3 = st.columns([1, 3, 0.6])

            with rcol1:
                pct_key = f"vc_pct_{rid}"
                pct_raw = st.text_input("Voucher %", value=st.session_state.get(pct_key, "10"),
                                        key=pct_key, placeholder="10")
                pct_clean = pct_raw.strip().replace("%", "")
                pct_val = int(pct_clean) if pct_clean.isdigit() else None
                if pct_raw and pct_val is None:
                    st.error("Whole number only, e.g. 10")

            with rcol2:
                st.markdown("**Eligible Remarks**")
                selected = []
                include_nr = False
                if not unique_remarks:
                    st.warning("No remarks found in the selected exclusion column.")
                else:
                    qc1, qc2, _ = st.columns([1, 1, 3])
                    with qc1:
                        if st.button("✅ Select All", key=f"vc_sa_{rid}"):
                            st.session_state[f"vc_remarks_{rid}"] = unique_remarks[:]
                    with qc2:
                        if st.button("❌ Clear All", key=f"vc_ca_{rid}"):
                            st.session_state[f"vc_remarks_{rid}"] = []

                    selected = st.multiselect(
                        "remarks", options=unique_remarks, default=[],
                        key=f"vc_remarks_{rid}", label_visibility="collapsed",
                        help="Only articles whose remark is selected here will be eligible for this voucher.",
                    )
                    include_nr = st.checkbox("Include blank/no-remark articles as eligible",
                                             value=False, key=f"vc_nr_{rid}")

                    eligible_set = set(selected)
                    if eligible_set or include_nr:
                        df_prev = zecom_df
                        if mp_col_for_preview and mp_col_for_preview in zecom_df.columns:
                            df_prev = zecom_df[zecom_df[mp_col_for_preview].astype(str)
                                               .str.strip().str.upper() == "YES"]
                        n_match = df_prev.iloc[:, excl_idx].astype(str).str.strip().isin(eligible_set).sum()
                        st.caption(f"📊 {n_match:,} articles with {marketplace}=YES match "
                                  f"(before price/stock filter).")
                    else:
                        st.caption("⚠️ No remarks selected and no-remark inclusion is off.")

            with rcol3:
                st.markdown("&nbsp;")
                remove_clicked = False
                if len(st.session_state.voucher_row_ids) > 1:
                    remove_clicked = st.button("🗑️", key=f"vc_rm_{rid}", help="Remove this voucher")

            return {"rid": rid, "pct": pct_val, "remarks": set(selected),
                   "include_no_remark": include_nr, "remove": remove_clicked}

        voucher_configs = []
        to_remove = None
        for pos, rid in enumerate(list(st.session_state.voucher_row_ids), start=1):
            row = _render_voucher_row(rid, pos)
            voucher_configs.append(row)
            if row["remove"]:
                to_remove = rid
            st.markdown("&nbsp;")

        if to_remove is not None:
            st.session_state.voucher_row_ids.remove(to_remove)
            st.rerun()

        if st.button("➕ Add Another Voucher"):
            new_id = st.session_state.voucher_row_counter
            st.session_state.voucher_row_ids.append(new_id)
            st.session_state.voucher_row_counter += 1
            st.rerun()

    # ── ⑤ GENERATE ────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("⑤ Generate")

    missing = []
    if not zecom_file:    missing.append("ZeCom Tracker")
    if not content_file:  missing.append("Content File")
    if not inv_file:      missing.append("Inventory File")
    if not mp_file:       missing.append(f"{marketplace} Export")

    if zecom_file:
        if not voucher_configs:
            missing.append("At least one voucher row")
        for i, row in enumerate(voucher_configs, start=1):
            if row["pct"] is None:
                missing.append(f"Voucher {i}: valid %")
            if not row["remarks"] and not row["include_no_remark"]:
                missing.append(f"Voucher {i}: at least one remark, or enable no-remark inclusion")

    if missing:
        st.info(f"Still needed: **{', '.join(missing)}**")

    if st.button("🚀 Generate Eligible SKU Lists", disabled=bool(missing), type="primary"):
        _run(zecom_file, content_file, inv_file, mp_file,
             region, marketplace, excl_idx, rrp_idx, srp_idx,
             special_articles, voucher_type, voucher_configs)

    # Always render last results (if any) — independent of the button click above,
    # so downloads/other widget interactions don't wipe them out.
    render_results()


# ─────────────────────────────────────────────────────────────────
# PROCESSING PIPELINE
# ─────────────────────────────────────────────────────────────────

def _run(zecom_file, content_file, inv_file, mp_file,
         region, marketplace, excl_idx, rrp_idx, srp_idx,
         special_articles, voucher_type, voucher_configs):

    with st.status("Processing…", expanded=True) as status:

        st.write(f"📊 ZeCom — {region}…")
        try:
            zecom_df = read_zecom(zecom_file.getvalue(), region)
            st.write(f"   ✓ {len(zecom_df):,} rows | excl=[{excl_idx}] rrp=[{rrp_idx}] srp=[{srp_idx}]")
        except Exception as e:
            st.error(f"ZeCom read error: {e}"); status.update(label="❌ Error", state="error"); return

        st.write("📦 Content file…")
        try:
            content_df = pd.read_excel(io.BytesIO(content_file.getvalue()), sheet_name="content")
            content_df = content_df[["Color_No", "EAN"]].dropna()
            content_df["EAN"]      = content_df["EAN"].astype(str).str.strip()
            content_df["Color_No"] = content_df["Color_No"].astype(str).str.strip()
            st.write(f"   ✓ {len(content_df):,} EAN mappings")
        except Exception as e:
            st.error(f"Content file error: {e}"); status.update(label="❌ Error", state="error"); return

        st.write("🏭 Inventory…")
        try:
            inv_df = read_inventory(inv_file.getvalue(), inv_file.name)
            if inv_df is None:
                st.error("Could not detect EAN/Stock columns.")
                status.update(label="❌ Error", state="error"); return
            st.write(f"   ✓ {len(inv_df):,} EANs | {(inv_df['Stock']>0).sum():,} in stock")
        except Exception as e:
            st.error(f"Inventory read error: {e}"); status.update(label="❌ Error", state="error"); return

        all_outputs = []   # one entry per voucher row

        for vi, row in enumerate(voucher_configs, start=1):
            pct = row["pct"]
            eligible_remarks  = row["remarks"]
            include_no_remark = row["include_no_remark"]

            st.write(f"⚙️  Voucher {vi}: **{pct}% {voucher_type}** — "
                     f"{len(eligible_remarks)} remark(s), "
                     f"{len(special_articles)} special exclusion(s)…")

            art = process_zecom(zecom_df, region, marketplace, excl_idx, rrp_idx, srp_idx,
                                eligible_remarks, include_no_remark, special_articles)
            n_elig  = (art["status"] == "eligible").sum()
            n_nr    = (art["status"] == "no_remark").sum()
            n_ineli = (art["status"] == "ineligible").sum()
            st.write(f"   ZeCom: {n_elig} eligible | {n_nr} no-remark | {n_ineli} ineligible")

            ean_df = map_to_eans(art, content_df, inv_df)
            n_ok   = len(eligible_ean_set(ean_df))
            st.write(f"   EANs in stock & eligible: {n_ok:,}")

            result = None
            pid_decisions = None

            if n_ok > 0:
                try:
                    if marketplace == "Lazada":
                        ids = process_lazada(ean_df, mp_file.getvalue())
                        st.write(f"   ✅ Lazada Shop SKUs → **{len(ids)}**")
                        result = {"mp": "Lazada", "ids": ids}
                    elif marketplace == "Shopee":
                        ids, pid_decisions = process_shopee(ean_df, mp_file.getvalue())
                        st.write(f"   ✅ Shopee Product IDs → **{len(ids)}** included "
                                 f"(of {len(pid_decisions)} total PIDs)")
                        result = {"mp": "Shopee", "ids": ids}
                    elif marketplace == "Zalora":
                        ann = process_zalora(ean_df, mp_file.getvalue(), content_df)
                        y  = (ann["Voucher Eligible"] == "Yes").sum()
                        nr = (ann["Voucher Eligible"] == "No Remark").sum()
                        st.write(f"   ✅ Zalora: {y} eligible | {nr} no-remark")
                        result = {"mp": "Zalora", "ann": ann, "yes_count": y}
                    elif marketplace == "TikTok":
                        ids, pid_decisions = process_tiktok(ean_df, mp_file.getvalue())
                        st.write(f"   ✅ TikTok Product IDs → **{len(ids)}** included "
                                 f"(of {len(pid_decisions)} total PIDs)")
                        result = {"mp": "TikTok", "ids": ids}
                except Exception as e:
                    st.error(f"   Error processing {marketplace}: {e}")
            else:
                st.warning("   No in-stock eligible EANs — marketplace output will be empty. "
                           "QC summary will still be generated.")

            summary_bytes = make_summary_excel(ean_df, region, marketplace, pct,
                                               voucher_type, pid_decisions)

            all_outputs.append({"pct": pct, "result": result,
                                "summary_bytes": summary_bytes, "pid_decisions": pid_decisions})

        status.update(label="✅ Done!", state="complete")

    # Persist results in session_state so they survive reruns
    # (e.g. clicking a download button triggers a rerun — without this,
    # the results would vanish since _run() wouldn't execute again).
    st.session_state["last_run"] = {
        "all_outputs":  all_outputs,
        "region":       region,
        "marketplace":  marketplace,
        "voucher_type": voucher_type,
        "generated_at": pd.Timestamp.now().strftime("%Y%m%d_%H%M%S"),
    }


def render_results():
    """Render the download section from session_state — survives reruns triggered by
    clicking a download button, switching tabs, etc. Only cleared on next Generate
    click or explicit 'Clear Results'."""
    last = st.session_state.get("last_run")
    if not last:
        return

    all_outputs  = last["all_outputs"]
    region       = last["region"]
    marketplace  = last["marketplace"]
    voucher_type = last["voucher_type"]
    today        = pd.Timestamp.now().strftime("%Y%m%d")
    vt_short     = "Bundle" if voucher_type == "Bundle Discount" else "VC"

    st.markdown("---")
    hcol1, hcol2 = st.columns([5, 1])
    with hcol1:
        st.subheader("⑥ Download Results")
        st.caption(f"From last Generate run ({marketplace} / {region}) — "
                  "stays available until you click Generate again.")
    with hcol2:
        if st.button("🧹 Clear"):
            del st.session_state["last_run"]
            st.rerun()

    for out in all_outputs:
        pct           = out["pct"]
        result        = out["result"]
        summary_bytes = out["summary_bytes"]
        pid_decisions = out["pid_decisions"]

        st.markdown(f"#### Voucher: {pct}% {voucher_type}")
        d1, d2 = st.columns(2)

        with d1:
            st.markdown("**Marketplace Output**")
            if result:
                mp = result["mp"]
                fname = f"{mp}_{region}_{pct}pct_{vt_short}_{today}.xlsx"
                if mp == "Zalora":
                    st.metric(f"{mp} — {pct}% {vt_short}", f"{result['yes_count']} eligible SKUs")
                    data = make_zalora_output(result["ann"])
                else:
                    label = "Shop SKUs" if mp == "Lazada" else "Product IDs"
                    st.metric(f"{mp} — {pct}% {vt_short}", f"{len(result['ids'])} {label}")
                    data = make_lazada_output(result["ids"]) if mp == "Lazada" else make_shopee_output(result["ids"])
                st.download_button(f"⬇️ Download {fname}", data=data, file_name=fname,
                                  mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  key=f"dl_main_{pct}")
            else:
                st.info("No marketplace output generated (0 eligible EANs).")

        with d2:
            st.markdown("**QC Summary (audit trail)**")
            summary_fname = f"QC_Summary_{marketplace}_{region}_{pct}pct_{vt_short}_{today}.xlsx"
            st.caption("Includes per-article/EAN status + reasons" +
                      (" + Product ID summary." if pid_decisions else "."))
            st.download_button(f"⬇️ Download {summary_fname}", data=summary_bytes,
                              file_name=summary_fname,
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              key=f"dl_summary_{pct}")

        st.markdown("---")


if __name__ == "__main__":
    main()
